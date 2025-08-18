import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os

def cancer_management_indicator(file_path: str, filename: str):
    """
    讀取 Excel 檔案並處理癌症管理指標。
    :param file_path: Excel 檔案路徑
    :return: new_file_path。
    """
    
    #讀取 Excel 檔案
    df = pd.read_excel(file_path)[1:]

    # 移除重複(條件:姓名和住院日和第一次介入實際熱量攝取及第一次建議熱量和第一次實際蛋白質攝取及第一次建議蛋白質一樣)
    filtered_df = df.drop_duplicates(subset=[
        "姓名", 
        "住院日", 
        "第一次介入實際熱量攝取", 
        "第一次建議熱量", 
        "第一次實際蛋白質攝取", 
        "第一次建議蛋白質"
    ]).reset_index(drop=True)

    # 熱量/建議 (%) = 最後一次介入實際熱量攝取 / 第一次建議熱量 * 100
    filtered_df["熱量/建議 (%)"] = filtered_df.apply(
        lambda row: row["最後一次介入實際熱量攝取"] / row["第一次建議熱量"] * 100
        if row["是否有營養會診單號"] == "●" and pd.notna(row["第一次建議熱量"]) and row["第一次建議熱量"] != 0
        else None,
        axis=1
    )
    # 蛋白質/建議 (%) = 最後一次實際蛋白質攝取 / 第一次建議蛋白質 * 100
    filtered_df["蛋白質/建議 (%)"] = filtered_df.apply(
        lambda row: row["最後一次實際蛋白質攝取"] / row["第一次建議蛋白質"] * 100
        if row["是否有營養會診單號"] == "●" and pd.notna(row["第一次建議蛋白質"]) and row["第一次建議蛋白質"] != 0
        else None,
        axis=1
    )
    # 補上「是否有營養會診單號」= "●" 的情況：
    # 若為空 且 第一次建議熱量有值，則代表應補「●」
    refill_mask = (filtered_df["是否有營養會診單號"].isna()) | (filtered_df["是否有營養會診單號"] == ' ') & (filtered_df["第一次建議熱量"].notna())
    refill_indices = filtered_df[refill_mask].index.tolist()
    filtered_df.loc[refill_indices, "是否有營養會診單號"] = "●"

    # 修正「訪視次數」：若為 0 且有第一次建議熱量 → 視為已訪視 → 改為 1
    mask = (filtered_df["訪視次數"] == 0) & (filtered_df["第一次建議熱量"].notna())
    corrected_indices = filtered_df[mask].index.tolist()
    filtered_df.loc[corrected_indices, "訪視次數"] = 1

    #中途檔
    output_path = './static/cancer_management_indicator/'+filename.split('.')[0]+'_已完成.xlsx'
    filtered_df.to_excel(output_path, index=False)

    #新增欄位篩選功能
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    ws.title = "2025Q1"
    max_row = ws.max_row
    max_col = ws.max_column
    last_col_letter = get_column_letter(max_col)
    table_range = f"A1:{last_col_letter}{max_row}"
    table = Table(displayName="DataTable", ref=table_range)
    ws.add_table(table)

    # 找出「訪視次數」欄位的 column index
    col_index = list(filtered_df.columns).index("訪視次數") + 1  # Excel從1開始
    # 套用紅色字體
    for row_idx in corrected_indices:
        excel_row = row_idx + 2  # DataFrame的第i列是Excel的第i+2列（含標題列）
        cell = ws.cell(row=excel_row, column=col_index)
        cell.font = Font(color="FF0000")  # 紅色

    wb.save(output_path)

    # 建立條件欄位（布林值）
    filtered_df["訪視次數大於2"] = filtered_df["訪視次數"] >= 2
    filtered_df["熱量達成 >80%"] = (
        (filtered_df["最後一次介入實際熱量攝取"] / filtered_df["第一次建議熱量"]) > 0.8
    ) & filtered_df["第一次建議熱量"].notna()
    filtered_df["蛋白質達成 >80%"] = (
        (filtered_df["最後一次實際蛋白質攝取"] / filtered_df["第一次實際蛋白質攝取"]) > 0.8
    ) & filtered_df["第一次實際蛋白質攝取"].notna()

    # 分組計算所需指標
    grouped = filtered_df.groupby("癌症團隊").agg(
        人次=('姓名', 'count'),
        有營養會診單=('是否有營養會診單號', lambda x: (x == '●').sum()),
        有會診人次=('第一次建議熱量', lambda x: x.notna().sum()),
        訪視人次總和=('訪視次數', 'sum'),
        訪視次數大於2筆數=('訪視次數大於2', 'sum'),
        熱量達成率高於80筆數=('熱量達成 >80%', 'sum'),
        蛋白質達成率高於80筆數=('蛋白質達成 >80%', 'sum')
    )

    # 計算百分比與會診完成率
    grouped["百分比 (%)"] = grouped.apply(
        lambda x: str(round(x["有營養會診單"] / x["人次"] * 100, 2)) + "%" if x["人次"] != 0 else "N/A", axis=1
    )
    grouped["會診完成率 (%)"] = grouped.apply(
        lambda x: str(round(x["有會診人次"] / x["有營養會診單"] * 100, 2)) + "%" if x["有營養會診單"] != 0 else "N/A", axis=1
    )
    grouped["熱量"] = grouped.apply(
        lambda x: str(round(x["熱量達成率高於80筆數"] / x["訪視次數大於2筆數"] * 100, 2)) + "%" if x["訪視次數大於2筆數"] != 0 else "N/A", axis=1
    )
    grouped["蛋白質"] = grouped.apply(
        lambda x: str(round(x["蛋白質達成率高於80筆數"] / x["訪視次數大於2筆數"] * 100, 2)) + "%" if x["訪視次數大於2筆數"] != 0 else "N/A", axis=1
    )


    # 調整欄位順序
    grouped = grouped[["人次", "有營養會診單", "百分比 (%)", "有會診人次", "會診完成率 (%)",
                    "訪視人次總和", "訪視次數大於2筆數", "熱量達成率高於80筆數", "蛋白質達成率高於80筆數", "熱量", "蛋白質"]]

    # 加上總計列
    summary_row = grouped[[
        "人次", "有營養會診單", "有會診人次",
        "訪視人次總和", "訪視次數大於2筆數",
        "熱量達成率高於80筆數", "蛋白質達成率高於80筆數"
    ]].sum()
    summary_row["百分比 (%)"] = summary_row["有營養會診單"] / summary_row["人次"] * 100
    summary_row["會診完成率 (%)"] = summary_row["有會診人次"] / summary_row["有營養會診單"] * 100
    summary_row["熱量"] = str(summary_row["熱量達成率高於80筆數"]/summary_row["訪視次數大於2筆數"]*100)+"%"
    summary_row["蛋白質"] = str(summary_row["蛋白質達成率高於80筆數"]/summary_row["訪視次數大於2筆數"]*100)+"%"
    summary_row["百分比 (%)"] = (
        str(round(summary_row["有營養會診單"] / summary_row["人次"] * 100, 2)) + "%"
        if summary_row["人次"] != 0 else "N/A"
    )
    summary_row["會診完成率 (%)"] = (
        str(round(summary_row["有會診人次"] / summary_row["有營養會診單"] * 100, 2)) + "%"
        if summary_row["有營養會診單"] != 0 else "N/A"
    )
    summary_row["熱量"] = (
        str(round(summary_row["熱量達成率高於80筆數"] / summary_row["訪視次數大於2筆數"] * 100, 2)) + "%"
        if summary_row["訪視次數大於2筆數"] != 0 else "N/A"
    )
    summary_row["蛋白質"] = (
        str(round(summary_row["蛋白質達成率高於80筆數"] / summary_row["訪視次數大於2筆數"] * 100, 2)) + "%"
        if summary_row["訪視次數大於2筆數"] != 0 else "N/A"
    )


    summary_row.name = "小計"
    grouped_with_total = pd.concat([grouped, pd.DataFrame([summary_row])])
    grouped_with_total = grouped_with_total.reset_index().rename(columns={"index": "癌症團隊"})

    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        grouped_with_total.to_excel(writer, index=False, sheet_name="年度會診完成率季報表")

    print(f"{output_path}已存檔")
    return output_path